import csv
from datetime import datetime
from decimal import Decimal, InvalidOperation
from io import BytesIO, TextIOWrapper

from django.contrib import messages
from django.db.models import Avg, Count, Max, Min
from django.http import HttpResponse, JsonResponse
from django.shortcuts import get_object_or_404, redirect, render
from django.utils import timezone
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill

from .forms import CSVUploadForm, SensorForm, SensorReadingForm
from .models import Sensor, SensorReading


def dashboard(request):
    stats = {
        'total_sensors': Sensor.objects.count(),
        'active_sensors': Sensor.objects.filter(is_active=True).count(),
        'total_readings': SensorReading.objects.count(),
        'critical_readings': SensorReading.objects.filter(status='critical').count(),
    }
    summary = SensorReading.objects.aggregate(
        avg_temperature=Avg('temperature'),
        avg_humidity=Avg('humidity'),
        avg_pressure=Avg('pressure'),
        min_battery=Min('battery_level'),
        max_temperature=Max('temperature'),
    )
    return render(request, 'sensors/dashboard.html', {
        'stats': stats,
        'summary': summary,
        'recent_readings': SensorReading.objects.select_related('sensor')[:10],
    })


def dashboard_data(request):
    latest = SensorReading.objects.select_related('sensor').order_by('-timestamp')[:50]
    location_summary = SensorReading.objects.values('sensor__location').annotate(
        readings=Count('id'), avg_temperature=Avg('temperature'), avg_humidity=Avg('humidity'), avg_battery=Avg('battery_level')
    ).order_by('sensor__location')
    status_summary = SensorReading.objects.values('status').annotate(total=Count('id')).order_by('status')
    timeline = list(reversed([{
        'timestamp': reading.timestamp.strftime('%Y-%m-%d %H:%M'),
        'sensor': reading.sensor.name,
        'temperature': float(reading.temperature),
        'humidity': float(reading.humidity),
        'pressure': float(reading.pressure),
        'battery_level': float(reading.battery_level),
        'status': reading.status,
    } for reading in latest]))
    return JsonResponse({
        'timeline': timeline,
        'location_summary': [{
            'location': item['sensor__location'],
            'readings': item['readings'],
            'avg_temperature': round(float(item['avg_temperature'] or 0), 2),
            'avg_humidity': round(float(item['avg_humidity'] or 0), 2),
            'avg_battery': round(float(item['avg_battery'] or 0), 2),
        } for item in location_summary],
        'status_summary': list(status_summary),
    })


def sensor_list(request):
    return render(request, 'sensors/sensor_list.html', {'sensors': Sensor.objects.annotate(reading_count=Count('readings'))})


def sensor_create(request):
    form = SensorForm(request.POST or None)
    if request.method == 'POST' and form.is_valid():
        form.save()
        messages.success(request, 'Sensor saved successfully.')
        return redirect('sensors:sensor_list')
    return render(request, 'sensors/sensor_form.html', {'form': form, 'title': 'Add Sensor'})


def sensor_update(request, pk):
    form = SensorForm(request.POST or None, instance=get_object_or_404(Sensor, pk=pk))
    if request.method == 'POST' and form.is_valid():
        form.save()
        messages.success(request, 'Sensor updated successfully.')
        return redirect('sensors:sensor_list')
    return render(request, 'sensors/sensor_form.html', {'form': form, 'title': 'Edit Sensor'})


def reading_list(request):
    readings = SensorReading.objects.select_related('sensor')
    sensor_id = request.GET.get('sensor')
    status = request.GET.get('status')
    if sensor_id:
        readings = readings.filter(sensor_id=sensor_id)
    if status:
        readings = readings.filter(status=status)
    return render(request, 'sensors/reading_list.html', {
        'readings': readings[:250], 'sensors': Sensor.objects.all(), 'selected_sensor': sensor_id, 'selected_status': status
    })


def reading_create(request):
    form = SensorReadingForm(request.POST or None)
    if request.method == 'POST' and form.is_valid():
        form.save()
        messages.success(request, 'Reading saved successfully.')
        return redirect('sensors:reading_list')
    return render(request, 'sensors/reading_form.html', {'form': form})


def reading_delete(request, pk):
    reading = get_object_or_404(SensorReading, pk=pk)
    if request.method == 'POST':
        reading.delete()
        messages.success(request, 'Reading deleted successfully.')
        return redirect('sensors:reading_list')
    return render(request, 'sensors/confirm_delete.html', {'reading': reading})


def upload_csv(request):
    form = CSVUploadForm(request.POST or None, request.FILES or None)
    if request.method == 'POST' and form.is_valid():
        uploaded = TextIOWrapper(request.FILES['csv_file'].file, encoding='utf-8')
        reader = csv.DictReader(uploaded)
        created = 0
        errors = []
        for line_number, row in enumerate(reader, start=2):
            try:
                sensor_name = row['sensor_name'].strip()
                location = row.get('location', 'Unknown').strip() or 'Unknown'
                sensor, _ = Sensor.objects.get_or_create(name=sensor_name, defaults={
                    'sensor_type': 'multi', 'location': location, 'installation_date': timezone.localdate()
                })
                SensorReading.objects.create(
                    sensor=sensor,
                    timestamp=parse_timestamp(row['timestamp']),
                    temperature=parse_decimal(row['temperature']),
                    humidity=parse_decimal(row['humidity']),
                    pressure=parse_decimal(row['pressure']),
                    battery_level=parse_decimal(row['battery_level']),
                    status=row.get('status', 'normal').strip() or 'normal',
                )
                created += 1
            except Exception as exc:
                errors.append(f'Line {line_number}: {exc}')
        if created:
            messages.success(request, f'{created} readings imported successfully.')
        for error in errors[:5]:
            messages.error(request, error)
        return redirect('sensors:reading_list')
    return render(request, 'sensors/upload_csv.html', {'form': form})


def parse_decimal(value):
    try:
        return Decimal(str(value).strip())
    except InvalidOperation as exc:
        raise ValueError(f'Invalid decimal value: {value}') from exc


def parse_timestamp(value):
    clean_value = str(value).strip()
    for fmt in ['%Y-%m-%d %H:%M:%S', '%Y-%m-%d %H:%M', '%Y-%m-%dT%H:%M:%S', '%Y-%m-%dT%H:%M']:
        try:
            return timezone.make_aware(datetime.strptime(clean_value, fmt))
        except ValueError:
            pass
    raise ValueError(f'Invalid timestamp: {value}')


def export_rows():
    return SensorReading.objects.select_related('sensor').order_by('-timestamp')


def export_csv(request):
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = 'attachment; filename="iot_sensor_readings.csv"'
    writer = csv.writer(response)
    writer.writerow(['sensor_name', 'sensor_type', 'location', 'timestamp', 'temperature', 'humidity', 'pressure', 'battery_level', 'status', 'notes'])
    for reading in export_rows():
        writer.writerow([reading.sensor.name, reading.sensor.sensor_type, reading.sensor.location, reading.timestamp.strftime('%Y-%m-%d %H:%M:%S'), reading.temperature, reading.humidity, reading.pressure, reading.battery_level, reading.status, reading.notes])
    return response


def export_excel(request):
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = 'IoT Sensor Readings'
    headers = ['sensor_name', 'sensor_type', 'location', 'timestamp', 'temperature', 'humidity', 'pressure', 'battery_level', 'status', 'notes']
    sheet.append(headers)
    for cell in sheet[1]:
        cell.font = Font(bold=True, color='FFFFFF')
        cell.fill = PatternFill('solid', fgColor='1F4E79')
    for reading in export_rows():
        sheet.append([reading.sensor.name, reading.sensor.sensor_type, reading.sensor.location, reading.timestamp.strftime('%Y-%m-%d %H:%M:%S'), float(reading.temperature), float(reading.humidity), float(reading.pressure), float(reading.battery_level), reading.status, reading.notes])
    output = BytesIO()
    workbook.save(output)
    output.seek(0)
    response = HttpResponse(output.read(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="iot_sensor_readings.xlsx"'
    return response
